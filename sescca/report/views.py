from django.shortcuts import render
from django.views.generic import CreateView
from openpyxl.styles.alignment import Alignment
from school.models import Student
from .models import Conduct, DailyData, WeeklyData
from django.http.response import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font

import pytz
from django.utils import timezone

import datetime


def convert_to_localtime(utctime):
    fmt = '%d/%m/%Y %H:%M'
    utc = utctime.replace(tzinfo=pytz.UTC)
    localtz = utc.astimezone(timezone.get_current_timezone())
    return localtz.strftime(fmt)

# Create your views here.
def IndividualReport(request, pk):
    student = Student.objects.get(id=pk)

    if request.method == "POST":
        start_date = request.POST.get("fromdate")
        finish_date = request.POST.get("todate")
        finish_date_search = datetime.datetime.strptime(finish_date, '%Y-%m-%d')
        finish_date_search += datetime.timedelta(days=1)
        finish_date_search = datetime.datetime.strftime(finish_date_search, '%Y-%m-%d')
        conduct_list = Conduct.objects.filter(student=student, created__range=[start_date, finish_date_search])
        daily_score_list = DailyData.objects.filter(student=student, created__range=[start_date, finish_date_search])
        weekly_score_list = WeeklyData.objects.filter(student=student, created__range=[start_date, finish_date_search])
        context = {'student':student, 'conduct_list':conduct_list, 'daily_score_list':daily_score_list, 'weekly_score_list':weekly_score_list}
        context['start_date'] = start_date
        context['finish_date'] = finish_date
        return render(request, 'school/student_detail.html', context)
    else:
        conduct_list = Conduct.objects.filter(student=student)
        daily_score_list = DailyData.objects.filter(student=student)
        weekly_score_list = WeeklyData.objects.filter(student=student)
        context = {'student':student, 'conduct_list':conduct_list, 'daily_score_list':daily_score_list, 'weekly_score_list':weekly_score_list}
        return render(request, 'school/student_detail.html', context)

class GenerateReport(CreateView):
    def get(self, request, *args, **kwargs):
        pk = self.kwargs.get(self.pk_url_kwarg)
        
        start_date = request.GET.get("start_date")
        finish_date = request.GET.get("finish_date")

        print(start_date, finish_date)

        student = Student.objects.get(id=pk)
        group = student.group_set.all()

        if not start_date:
            conduct_list = Conduct.objects.filter(student=student)
            daily_score_list = DailyData.objects.filter(student=student)
            weekly_score_list = WeeklyData.objects.filter(student=student)
        else: 
            finish_date_search = datetime.datetime.strptime(finish_date, '%Y-%m-%d')
            finish_date_search += datetime.timedelta(days=1)
            finish_date_search = datetime.datetime.strftime(finish_date_search, '%Y-%m-%d')
            conduct_list = Conduct.objects.filter(student=student, created__range=[start_date, finish_date_search])
            daily_score_list = DailyData.objects.filter(student=student, created__range=[start_date, finish_date_search])
            weekly_score_list = WeeklyData.objects.filter(student=student, created__range=[start_date, finish_date_search])

        # Cabecera
        wb = Workbook()
        wb.encoding = 'utf-8'
        ws = wb.active
        title = ws['B2']
        title.value = 'REPORTE DE ESTUDIANTE'
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('B2:M3')
        name = ws['B4']
        name.font = Font(name='Calibri', size="12")
        name.value = 'Estudiante: ' + str(student.name) + ' ' + str(student.last_name)
        name.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('B4:M5')
        ws['B6'] = 'SEDE'
        ws.merge_cells('B6:D6')
        ws['E6'] = 'JORNADA'
        ws.merge_cells('E6:G6')
        ws['H6'] = 'CURSO'
        ws.merge_cells('H6:J6')
        ws['K6'] = 'GRUPO'
        ws.merge_cells('K6:M6')
        cont = 7
        ws.cell(row=cont, column=2).value = student.campus.name
        ws.merge_cells('B{}:D{}'.format(cont, cont))
        ws.cell(row=cont, column=5).value = student.worktime.name
        ws.merge_cells('E{}:G{}'.format(cont, cont))
        ws.cell(row=cont, column=8).value = str(student.section.grade) + student.section.letter
        ws.merge_cells('H{}:J{}'.format(cont, cont))
        if group:
            ws.cell(row=cont, column=11).value = group[0]
        else:
            ws.cell(row=cont, column=11).value = 'No asignado'
        ws.merge_cells('K{}:M{}'.format(cont, cont))

        # Cuerpo
        subtitle1 = ws['B8']
        subtitle1.value = 'CONDUCTAS'
        subtitle1.alignment = Alignment(horizontal='left', vertical='center')
        subtitle1.font = Font(name='Calibri', size="12")
        ws.merge_cells('B8:M9')

        row = 10
        if not conduct_list:
            cell = ws['B{}'.format(row)]
            cell.value = 'No hay registros'
            cell.font = Font(name='Calibri', size="12")
            ws.merge_cells('B{}:M{}'.format(row, row))
            row += 1
        else:
            date1 = ws['B{}'.format(row)]
            date1.value = 'Fecha'
            date1.alignment = Alignment(horizontal='center', vertical='center')
            date1.font = Font(name='Calibri', size="10")
            ws.merge_cells('B{}:G{}'.format(row, row))

            detail1 = ws['H{}'.format(row)]
            detail1.value = 'Conducta'
            detail1.alignment = Alignment(horizontal='center', vertical='center')
            detail1.font = Font(name='Calibri', size="10")
            ws.merge_cells('H{}:M{}'.format(row, row))

            row += 1

            for conduct in conduct_list:
                cell_date = ws['B{}'.format(row)]
                cell_date.value = convert_to_localtime(conduct.created)
                cell_date.alignment = Alignment(horizontal='center', vertical='center')
                cell_date.font = Font(name='Calibri', size="12")
                ws.merge_cells('B{}:G{}'.format(row, row))

                cell_name = ws['H{}'.format(row)]
                cell_name.value = conduct.conduct
                cell_name.alignment = Alignment(horizontal='center', vertical='center')
                cell_name.font = Font(name='Calibri', size="12")
                ws.merge_cells('H{}:M{}'.format(row, row))
            
                row += 1

        subtitle2 = ws['B{}'.format(row)]
        subtitle2.value = 'PUNTAJES DIARIOS'
        subtitle2.alignment = Alignment(horizontal='left', vertical='center')
        subtitle2.font = Font(name='Calibri', size="12")
        ws.merge_cells('B{}:M{}'.format(row, row+1))
        row += 2


        if not daily_score_list:
            cell = ws['B{}'.format(row)]
            cell.value = 'No hay registros'
            cell.font = Font(name='Calibri', size="12")
            ws.merge_cells('B{}:M{}'.format(row, row))
            row += 1
        else:
            date1 = ws['B{}'.format(row)]
            date1.value = 'Fecha'
            date1.alignment = Alignment(horizontal='center', vertical='center')
            date1.font = Font(name='Calibri', size="10")
            ws.merge_cells('B{}:G{}'.format(row, row))

            detail1 = ws['H{}'.format(row)]
            detail1.value = 'Puntaje'
            detail1.alignment = Alignment(horizontal='center', vertical='center')
            detail1.font = Font(name='Calibri', size="10")
            ws.merge_cells('H{}:M{}'.format(row, row))

            row += 1

            for daily_score in daily_score_list:
                cell_date = ws['B{}'.format(row)]
                cell_date.value = convert_to_localtime(daily_score.updated)
                cell_date.alignment = Alignment(horizontal='center', vertical='center')
                cell_date.font = Font(name='Calibri', size="12")
                ws.merge_cells('B{}:G{}'.format(row, row))

                cell_name = ws['H{}'.format(row)]
                cell_name.value = daily_score.daily_score
                cell_name.alignment = Alignment(horizontal='center', vertical='center')
                cell_name.font = Font(name='Calibri', size="12")
                ws.merge_cells('H{}:M{}'.format(row, row))
            
                row += 1

        subtitle3 = ws['B{}'.format(row)]
        subtitle3.value = 'PUNTAJES SEMANALES'
        subtitle3.alignment = Alignment(horizontal='left', vertical='center')
        subtitle3.font = Font(name='Calibri', size="12")
        ws.merge_cells('B{}:M{}'.format(row, row+1))
        row += 2
        

        if not weekly_score_list:
            cell = ws['B{}'.format(row)]
            cell.value = 'No hay registros'
            cell.font = Font(name='Calibri', size="12")
            ws.merge_cells('B{}:M{}'.format(row, row))
            row += 1
        else:
            date1 = ws['B{}'.format(row)]
            date1.value = 'Fecha'
            date1.alignment = Alignment(horizontal='center', vertical='center')
            date1.font = Font(name='Calibri', size="10")
            ws.merge_cells('B{}:G{}'.format(row, row))

            detail1 = ws['H{}'.format(row)]
            detail1.value = 'Puntaje semana'
            detail1.alignment = Alignment(horizontal='center', vertical='center')
            detail1.font = Font(name='Calibri', size="10")
            ws.merge_cells('H{}:M{}'.format(row, row))

            row += 1

            for weekly_score in weekly_score_list:
                cell_date = ws['B{}'.format(row)]
                cell_date.value = convert_to_localtime(weekly_score.updated)
                cell_date.alignment = Alignment(horizontal='center', vertical='center')
                cell_date.font = Font(name='Calibri', size="12")
                ws.merge_cells('B{}:G{}'.format(row, row))

                cell_name = ws['H{}'.format(row)]
                cell_name.value = weekly_score.weekly_score
                cell_name.alignment = Alignment(horizontal='center', vertical='center')
                cell_name.font = Font(name='Calibri', size="12")
                ws.merge_cells('H{}:M{}'.format(row, row))
            
                row += 1

        # Descarga de Archivo
        nombre_archivo = student.name + '_Report.xlsx'

        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)

        response["Content-Disposition"] = contenido

        wb.save(response)

        return response

